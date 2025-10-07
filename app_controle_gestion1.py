import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
from scipy.optimize import linprog
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_absolute_error
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from io import BytesIO
from openpyxl import Workbook

# Configuration de la page
st.set_page_config(
    page_title="Contrôle de Gestion",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personnalisé
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .section-header {
        font-size: 1.8rem;
        color: #2e86ab;
        border-bottom: 2px solid #2e86ab;
        padding-bottom: 0.5rem;
        margin-top: 2rem;
    }
    .formula-box {
        background-color: #f0f8ff;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #1f77b4;
        margin: 1rem 0;
    }
    .exercise-box {
        background-color: #fffaf0;
        padding: 1.5rem;
        border-radius: 10px;
        border: 2px solid #ffa500;
        margin: 1rem 0;
    }
    .footer {
        background-color: #f8f9fa;
        padding: 2rem;
        border-radius: 10px;
        margin-top: 3rem;
        border-top: 3px solid #1f77b4;
    }
    .author-info {
        text-align: center;
        padding: 1rem;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border-radius: 10px;
        margin-top: 2rem;
    }
</style>
""", unsafe_allow_html=True)

# Menu de navigation
def main():
    st.sidebar.title("📊 Navigation")
    sections = {
        "🏠 Accueil": "accueil",
        "📚 Introduction": "introduction",
        "💰 Gestion Budgétaire": "gestion_budgetaire",
        "📈 Budgets Opérationnels": "budgets_operationnels",
        "🏭 Budget des Investissements": "budget_investissements",
        "💳 Budget de Trésorerie": "budget_tresorerie",
        "📊 Contrôle Budgétaire": "controle_budgetaire",
        "🎯 Études de Cas": "etudes_cas",
        "🤖 Applications Pratiques": "applications_pratiques",
        "📖 Ressources": "ressources"
    }
    
    selection = st.sidebar.radio("Sections", list(sections.keys()))
    
    # Affichage de la section sélectionnée
    if sections[selection] == "accueil":
        accueil()
    elif sections[selection] == "introduction":
        introduction()
    elif sections[selection] == "gestion_budgetaire":
        gestion_budgetaire()
    elif sections[selection] == "budgets_operationnels":
        budgets_operationnels()
    elif sections[selection] == "budget_investissements":
        budget_investissements()
    elif sections[selection] == "budget_tresorerie":
        budget_tresorerie()
    elif sections[selection] == "controle_budgetaire":
        controle_budgetaire()
    elif sections[selection] == "etudes_cas":
        etudes_cas()
    elif sections[selection] == "applications_pratiques":
        applications_pratiques()
    elif sections[selection] == "ressources":
        ressources()
    
    # Ajouter le footer à toutes les pages
    footer()

# Section Accueil
def accueil():
    st.markdown('<div class="main-header">Bienvenue dans l\'application Contrôle de Gestion</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.write("""
        ### Objectifs de l'Application
        
        Cette application interactive vous permet de :
        
        - ✅ Comprendre les concepts fondamentaux du contrôle de gestion
        - 📈 Maîtriser les méthodes de budgétisation et de prévision
        - 🧮 Appliquer les formules essentielles via des calculateurs interactifs
        - 🎯 Résoudre des exercices pratiques avec correction détaillée
        - 📊 Visualiser les données et résultats grâce à des graphiques interactifs
        
        ### Public Visé
        
        - **Étudiants** en gestion, finance ou comptabilité
        - **Professionnels** souhaitant réviser ou approfondir leurs connaissances
        - **Entrepreneurs** désireux de mieux piloter leur entreprise
        """)
    
    with col2:
        st.image("https://cdn.pixabay.com/photo/2017/09/07/08/54/money-2724241_1280.jpg", 
                caption="Piloter la performance de l'entreprise")
    
    # Schéma du cycle de contrôle de gestion
    st.markdown("### 🔄 Cycle du Contrôle de Gestion")
    
    fig = go.Figure(go.Indicator(
        mode = "gauge+number+delta",
        value = 25,
        domain = {'x': [0, 1], 'y': [0, 1]},
        title = {'text': "Cycle de Contrôle"},
        delta = {'reference': 100},
        gauge = {
            'axis': {'range': [None, 100]},
            'steps': [
                {'range': [0, 25], 'color': "lightgray"},
                {'range': [25, 50], 'color': "gray"},
                {'range': [50, 75], 'color': "darkgray"},
                {'range': [75, 100], 'color': "black"}
            ],
            'bar': {'color': "darkblue"},
        }
    ))
    
    fig.update_layout(
        height=300,
        annotations=[
            dict(
                x=0.13, y=0.5,
                xref="paper", yref="paper",
                text="1. Planification",
                showarrow=False,
                font=dict(size=12)
            ),
            dict(
                x=0.38, y=0.85,
                xref="paper", yref="paper",
                text="2. Budgétisation",
                showarrow=False,
                font=dict(size=12)
            ),
            dict(
                x=0.63, y=0.5,
                xref="paper", yref="paper",
                text="3. Contrôle",
                showarrow=False,
                font=dict(size=12)
            ),
            dict(
                x=0.38, y=0.15,
                xref="paper", yref="paper",
                text="4. Analyse",
                showarrow=False,
                font=dict(size=12)
            )
        ]
    )
    
    st.plotly_chart(fig, use_container_width=True, key="cycle_controle")

# Section Introduction
def introduction():
    st.markdown('<div class="section-header">📚 Introduction au Contrôle de Gestion</div>', unsafe_allow_html=True)
    
    tab1, tab2, tab3, tab4 = st.tabs(["🎯 Définition", "👨‍💼 Rôles", "💪 Compétences", "🧠 Quiz"])
    
    with tab1:
        st.write("""
        ### Qu'est-ce que le Contrôle de Gestion ?
        
        Le contrôle de gestion est un **processus de pilotage** qui permet à l'entreprise d'atteindre ses objectifs 
        stratégiques en optimisant l'utilisation de ses ressources.
        
        #### Enjeux Principaux :
        - 🎯 **Piloter la performance** globale de l'organisation
        - 📊 **Mesurer et analyser** les résultats
        - 🔍 **Identifier les écarts** entre prévisions et réalisations
        - 🛠️ **Proposer des actions correctives**
        - 🔮 **Anticiper** l'avenir via la prévision
        """)
    
    with tab2:
        st.write("""
        ### Rôles et Missions du Contrôleur de Gestion
        
        Le contrôleur de gestion est un **partenaire de gestion** ("business partner") :
        
        #### Missions Principales :
        1. **Participer à la définition des objectifs stratégiques**
        2. **Élaborer les budgets prévisionnels**
        3. **Contrôler et analyser les résultats**
        4. **Calculer et analyser les écarts**
        5. **Proposer des actions correctives**
        6. **Anticiper les problèmes budgétaires**
        7. **Créer des outils de pilotage** (tableaux de bord)
        8. **Établir des prévisions financières** (3-5 ans)
        """)
        
        # Exemple interactif de rôle
        role_selection = st.selectbox(
            "Découvrez un rôle en détail :",
            ["Sélectionnez un rôle", "Business Partner", "Analyste des écarts", "Créateur de tableaux de bord"]
        )
        
        if role_selection == "Business Partner":
            st.info("""
            **Business Partner** : 
            - Travaille en étroite collaboration avec les responsables opérationnels
            - Traduit la stratégie en indicateurs mesurables
            - Accompagne la prise de décision avec des analyses financières
            """)
    
    with tab3:
        st.write("""
        ### Compétences Requises
        
        #### Compétences Techniques :
        - 📊 **Comptabilité et finance** avancées
        - ⚖️ **Connaissance législative** (lois financières)
        - 💻 **Maîtrise des outils informatiques** (Excel, ERP, Power BI, Tableau)
        - 🏢 **Connaissance de l'organigramme** et des métiers de l'entreprise
        - 🔍 **Capacité d'analyse** des données et des chiffres
        
        #### Compétences Comportementales :
        - 🗣️ **Communication** et pédagogie
        - 🤝 **Collaboration** avec les équipes opérationnelles
        - 💡 **Esprit d'analyse** et de synthèse
        - 🔮 **Vision stratégique**
        """)
        
        # Diagramme de compétences interactif
        skills = {
            'Comptabilité/Finance': 95,
            'Analyse de données': 90,
            'Outils informatiques': 85,
            'Communication': 80,
            'Connaissance métier': 75
        }
        
        fig = go.Figure(data=[
            go.Bar(
                x=list(skills.values()),
                y=list(skills.keys()),
                orientation='h',
                marker_color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd']
            )
        ])
        
        fig.update_layout(
            title="Niveau de Compétences Requises",
            xaxis_title="Niveau (%)",
            yaxis_title="Compétences",
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True, key="competences_graph")
    
    with tab4:
        st.write("### 🧠 Quiz d'Évaluation des Connaissances")
        
        quiz_questions = {
            "Quel est le rôle principal du contrôle de gestion ?": {
                "options": ["Contrôler les employés", "Piloter la performance", "Faire la comptabilité", "Vérifier les stocks"],
                "correct": 1
            },
            "Le contrôleur de gestion travaille principalement avec :": {
                "options": ["Les clients uniquement", "La direction générale et les opérationnels", "Les fournisseurs", "Les banques"],
                "correct": 1
            },
            "Quelle compétence est la moins essentielle pour un contrôleur de gestion ?": {
                "options": ["Analyse financière", "Programmation informatique", "Communication", "Connaissance des processus métier"],
                "correct": 1
            }
        }
        
        score = 0
        user_answers = {}
        
        for i, (question, data) in enumerate(quiz_questions.items()):
            user_answers[i] = st.radio(question, data["options"], key=f"q{i}")
            if user_answers[i] == data["options"][data["correct"]]:
                score += 1
        
        if st.button("📊 Voir mon score"):
            st.success(f"Score : {score}/{len(quiz_questions)}")
            if score == len(quiz_questions):
                st.balloons()
                st.success("🎉 Excellent ! Vous maîtrisez les bases du contrôle de gestion.")

# Section Gestion Budgétaire
def gestion_budgetaire():
    st.markdown('<div class="section-header">💰 Gestion Budgétaire</div>', unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["📖 Théorie", "🔄 Processus", "🧮 Exercice"])
    
    with tab1:
        st.write("""
        ### Pourquoi la Gestion Budgétaire ?
        
        #### Contexte Historique :
        - Passage d'une **logique production** à une **logique marketing**
        - Émergence des **coûts préétablis**
        - Augmentation de la taille des entreprises et **décentralisation**
        
        #### Définition :
        La gestion budgétaire est un **mode de gestion à court terme** qui englobe tous les aspects 
        de l'activité dans un ensemble cohérent de **prévisions chiffrées**.
        
        #### Objectifs :
        - ✅ **Coordonner** les actions de l'entreprise
        - 📊 **Anticiper** les besoins en ressources
        - 🎯 **Déléguer** les responsabilités
        - 🔍 **Contrôler** la performance
        - 🛠️ **Corriger** les écarts
        """)
    
    with tab2:
        st.write("""
        ### Le Processus de Planification
        
        #### Les 3 Niveaux de Planification :
        """)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.info("""
            **🎯 Plan Stratégique** (3-5 ans)
            - Vision et objectifs long terme
            - Choix des marchés et technologies
            - Moyens (croissance interne/externe)
            """)
        
        with col2:
            st.warning("""
            **📋 Plans Opérationnels** (1-3 ans)
            - Déclinaison de la stratégie
            - Par fonction/département
            - Plan d'investissement, de financement
            """)
        
        with col3:
            st.success("""
            **💰 Budgétisation** (1 an)
            - Chiffrage détaillé
            - Budgets opérationnels
            - Contrôle mensuel
            """)
        
        # Schéma interactif du cycle
        st.write("### 🔄 Cycle de Pilotage")
        
        cycle_data = {
            'Étape': ['Planification', 'Budgétisation', 'Réalisation', 'Contrôle', 'Analyse', 'Actions Correctives'],
            'Description': [
                'Définir les objectifs',
                'Chiffrer les plans',
                'Exécuter les activités',
                'Comparer réel vs budget',
                'Comprendre les écarts',
                'Ajuster et améliorer'
            ]
        }
        
        df_cycle = pd.DataFrame(cycle_data)
        
        fig = px.line_polar(df_cycle, r=[1]*6, theta='Étape', line_close=True)
        fig.update_traces(fill='toself')
        fig.update_layout(
            polar=dict(
                radialaxis=dict(visible=True, range=[0, 1])
            ),
            showlegend=False,
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True, key="cycle_pilotage")
    
    with tab3:
        st.markdown('<div class="exercise-box">🧮 Exercice : Budget des Ventes d\'une PME</div>', unsafe_allow_html=True)
        
        st.write("""
        **Énoncé :**
        Une PME souhaite établir son budget des ventes pour l'année N+1. 
        Les ventes des 5 dernières années sont les suivantes :
        """)
        
        # Données de l'exercice
        data = {
            'Année': [2019, 2020, 2021, 2022, 2023],
            'Ventes (k€)': [120, 135, 158, 172, 190]
        }
        
        df = pd.DataFrame(data)
        st.dataframe(df, use_container_width=True)
        
        st.write("""
        **Questions :**
        1. Calculez le taux de croissance annuel moyen
        2. Projetez les ventes pour 2024 avec une croissance de 8%
        3. Établissez une répartition trimestrielle (Q1:20%, Q2:25%, Q3:30%, Q4:25%)
        """)
        
        # Zone de réponse
        with st.form("exercice_ventes"):
            croissance_moyenne = st.number_input("1. Taux de croissance annuel moyen (%)", min_value=0.0, max_value=50.0, value=0.0)
            ventes_2024 = st.number_input("2. Ventes projetées 2024 (k€)", min_value=0, value=0)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                q1 = st.number_input("Q1 (k€)", min_value=0, value=0)
            with col2:
                q2 = st.number_input("Q2 (k€)", min_value=0, value=0)
            with col3:
                q3 = st.number_input("Q3 (k€)", min_value=0, value=0)
            with col4:
                q4 = st.number_input("Q4 (k€)", min_value=0, value=0)
            
            submitted = st.form_submit_button("📤 Valider mes réponses")
            
            if submitted:
                # Correction
                croissance_correcte = ((190/120)**(1/4)-1)*100
                ventes_correctes = 190 * 1.08
                repartition_correcte = {
                    'Q1': ventes_correctes * 0.20,
                    'Q2': ventes_correctes * 0.25,
                    'Q3': ventes_correctes * 0.30,
                    'Q4': ventes_correctes * 0.25
                }
                
                st.write("### 📝 Correction")
                
                col_c1, col_c2 = st.columns(2)
                with col_c1:
                    st.write(f"**1. Croissance moyenne :** {croissance_correcte:.2f}%")
                    st.write(f"**2. Ventes 2024 :** {ventes_correctes:.0f} k€")
                with col_c2:
                    st.write("**3. Répartition trimestrielle :**")
                    for trim, valeur in repartition_correcte.items():
                        st.write(f"- {trim} : {valeur:.0f} k€")

# Section Budgets Opérationnels
def budgets_operationnels():
    st.markdown('<div class="section-header">📈 Budgets Opérationnels</div>', unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["📈 Ventes", "🏭 Production", "📦 Approvisionnements"])
    
    with tab1:
        st.write("### Budget des Ventes : Méthode des Moindres Carrés")
        
        st.markdown("""
        <div class="formula-box">
        <strong>Formule :</strong> y = ax + b<br>
        Où :<br>
        a = [ Σ(xi - x̄)(yi - ȳ) ] / Σ(xi - x̄)²<br>
        b = ȳ - a * x̄
        </div>
        """, unsafe_allow_html=True)
        
        # Calculateur interactif
        st.subheader("🧮 Calculateur de Droite de Tendance")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("**Entrez vos données :**")
            periodes_input = st.text_input("Périodes (séparées par des virgules)", "1,2,3,4,5")
            ventes_input = st.text_input("Ventes (séparées par des virgules)", "100,120,150,170,190")
        
        with col2:
            st.write("**Paramètres :**")
            periode_prevision = st.number_input("Période à prévoir", min_value=1, value=6)
            confidence = st.slider("Intervalle de confiance (%)", 80, 95, 90)
        
        if st.button("📊 Calculer la prévision"):
            try:
                # Conversion des données
                x = np.array([float(x.strip()) for x in periodes_input.split(',')])
                y = np.array([float(y.strip()) for y in ventes_input.split(',')])
                
                # Calcul des coefficients
                x_mean = np.mean(x)
                y_mean = np.mean(y)
                a = np.sum((x - x_mean) * (y - y_mean)) / np.sum((x - x_mean) ** 2)
                b = y_mean - a * x_mean
                
                # Affichage des résultats
                col_r1, col_r2 = st.columns(2)
                
                with col_r1:
                    st.metric("Coefficient a", f"{a:.2f}")
                    st.metric("Coefficient b", f"{b:.2f}")
                    prevision = a * periode_prevision + b
                    st.metric(f"Prévision période {periode_prevision}", f"{prevision:.0f}")
                
                with col_r2:
                    st.latex(f"y = {a:.2f}x + {b:.2f}")
                
                # Graphique
                fig, ax = plt.subplots(figsize=(10, 6))
                ax.scatter(x, y, color='blue', label='Données historiques', s=50)
                
                # Droite de tendance
                x_trend = np.linspace(min(x), max(x)+1, 100)
                y_trend = a * x_trend + b
                ax.plot(x_trend, y_trend, 'r-', label='Droite de tendance', linewidth=2)
                
                # Point de prévision
                ax.scatter([periode_prevision], [prevision], color='green', s=100, label='Prévision', marker='*')
                
                ax.set_xlabel('Périodes')
                ax.set_ylabel('Ventes')
                ax.legend()
                ax.grid(True, alpha=0.3)
                
                st.pyplot(fig)
                
            except Exception as e:
                st.error(f"Erreur dans les données : {e}")
    
    with tab2:
        st.write("### Budget de Production : Programmation Linéaire")
        
        st.markdown("""
        <div class="formula-box">
        <strong>Problème type :</strong> Maximiser Z = Marge1 × x + Marge2 × y<br>
        <strong>Sous contraintes :</strong><br>
        - Contraintes techniques : a₁x + b₁y ≤ C₁<br>
        - Contraintes marché : x ≤ Dmax₁, y ≤ Dmax₂<br>
        - Contraintes logiques : x ≥ 0, y ≥ 0
        </div>
        """, unsafe_allow_html=True)
        
        # Exercice de programmation linéaire
        st.markdown('<div class="exercise-box">🧮 Exercice : Optimisation de la Production</div>', unsafe_allow_html=True)
        
        st.write("""
        **Énoncé :**
        Une entreprise fabrique 2 produits P1 et P2 avec les caractéristiques suivantes :
        """)
        
        col_data1, col_data2 = st.columns(2)
        
        with col_data1:
            st.write("**Produit P1 :**")
            marge1 = st.number_input("Marge unitaire P1 (€)", min_value=0, value=50)
            temps_usinage1 = st.number_input("Temps usinage P1 (h)", min_value=0.0, value=2.0)
            temps_montage1 = st.number_input("Temps montage P1 (h)", min_value=0.0, value=1.0)
            demande1 = st.number_input("Demande max P1", min_value=0, value=100)
        
        with col_data2:
            st.write("**Produit P2 :**")
            marge2 = st.number_input("Marge unitaire P2 (€)", min_value=0, value=70)
            temps_usinage2 = st.number_input("Temps usinage P2 (h)", min_value=0.0, value=1.0)
            temps_montage2 = st.number_input("Temps montage P2 (h)", min_value=0.0, value=2.0)
            demande2 = st.number_input("Demande max P2", min_value=0, value=80)
        
        st.write("**Contraintes de capacité :**")
        cap_usinage = st.number_input("Capacité usinage (h)", min_value=0, value=200)
        cap_montage = st.number_input("Capacité montage (h)", min_value=0, value=180)
        
        if st.button("🔍 Optimiser la production"):
            # Résolution du problème
            c = [-marge1, -marge2]  # On minimise l'opposé de la marge
            A = [
                [temps_usinage1, temps_usinage2],
                [temps_montage1, temps_montage2]
            ]
            b = [cap_usinage, cap_montage]
            x_bounds = [(0, demande1), (0, demande2)]
            
            result = linprog(c, A_ub=A, b_ub=b, bounds=x_bounds, method='highs')
            
            if result.success:
                q1_opt, q2_opt = result.x
                marge_totale = -result.fun
                
                st.success("✅ Solution optimale trouvée !")
                
                col_res1, col_res2, col_res3 = st.columns(3)
                with col_res1:
                    st.metric("Quantité P1 optimale", f"{q1_opt:.0f}")
                with col_res2:
                    st.metric("Quantité P2 optimale", f"{q2_opt:.0f}")
                with col_res3:
                    st.metric("Marge totale optimale", f"{marge_totale:.0f} €")
                
                # Vérification des contraintes
                st.write("**Vérification des contraintes :**")
                col_ver1, col_ver2 = st.columns(2)
                with col_ver1:
                    usage_usinage = temps_usinage1 * q1_opt + temps_usinage2 * q2_opt
                    st.write(f"Usinage : {usage_usinage:.1f}h / {cap_usinage}h")
                with col_ver2:
                    usage_montage = temps_montage1 * q1_opt + temps_montage2 * q2_opt
                    st.write(f"Montage : {usage_montage:.1f}h / {cap_montage}h")
    
    with tab3:
        st.write("### Budget des Approvisionnements : Modèle de Wilson")
        
        st.markdown("""
        <div class="formula-box">
        <strong>Formule du lot économique :</strong><br>
        Q* = √(2 × D × CL / p)<br>
        Où :<br>
        D = Consommation annuelle<br>
        CL = Coût de lancement d'une commande<br>
        p = Coût de possession unitaire annuel
        </div>
        """, unsafe_allow_html=True)
        
        # Calculateur de Wilson
        st.subheader("🧮 Calculateur du Lot Économique")
        
        col_w1, col_w2 = st.columns(2)
        
        with col_w1:
            consommation = st.number_input("Consommation annuelle (unités)", min_value=1, value=10000)
            cout_lancement = st.number_input("Coût de lancement (€/commande)", min_value=1.0, value=150.0)
        
        with col_w2:
            prix_unitaire = st.number_input("Prix unitaire (€)", min_value=0.1, value=25.0)
            taux_possession = st.slider("Taux de possession annuel (%)", 1, 30, 15) / 100
        
        if st.button("📦 Calculer le lot économique"):
            cout_possession = prix_unitaire * taux_possession
            lot_eco = np.sqrt(2 * consommation * cout_lancement / cout_possession)
            nbr_commandes = consommation / lot_eco
            periode_eco = 365 / nbr_commandes
            
            st.success("📊 Résultats du modèle de Wilson")
            
            col_res1, col_res2, col_res3 = st.columns(3)
            with col_res1:
                st.metric("Lot économique", f"{lot_eco:.0f} unités")
            with col_res2:
                st.metric("Nombre de commandes/an", f"{nbr_commandes:.1f}")
            with col_res3:
                st.metric("Période économique", f"{periode_eco:.0f} jours")
            
            # Graphique des coûts
            quantites = np.linspace(lot_eco * 0.5, lot_eco * 1.5, 50)
            couts_lancement = (consommation / quantites) * cout_lancement
            couts_possession = (quantites / 2) * cout_possession
            couts_totaux = couts_lancement + couts_possession
            
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.plot(quantites, couts_lancement, 'b-', label='Coût de lancement', linewidth=2)
            ax.plot(quantites, couts_possession, 'r-', label='Coût de possession', linewidth=2)
            ax.plot(quantites, couts_totaux, 'g-', label='Coût total', linewidth=3)
            ax.axvline(lot_eco, color='orange', linestyle='--', label=f'Lot économique = {lot_eco:.0f}')
            ax.set_xlabel('Quantité commandée')
            ax.set_ylabel('Coût (€)')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            st.pyplot(fig)

# Section Budget des Investissements
def budget_investissements():
    st.markdown('<div class="section-header">🏭 Budget des Investissements</div>', unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["📖 Théorie", "🧮 VAN", "🔍 Comparaison"])
    
    with tab1:
        st.write("""
        ### Évaluation des Investissements
        
        #### Méthodes d'Évaluation :
        
        **1. Délai de Récupération (Payback)**
        - Temps nécessaire pour récupérer le capital investi
        - Avantage : Simple et intuitif
        - Limite : Ignore la valeur temps de l'argent
        
        **2. Taux de Rentabilité Comptable (TRC)**
        - TRC = Bénéfice Annuel Moyen / Investissement Initial
        - Avantage : Facile à calculer
        - Limite : Basé sur des données comptables
        
        **3. Valeur Actuelle Nette (VAN)**
        - VAN = Σ(Flux actualisés) - Investissement Initial
        - Avantage : Prend en compte la valeur temps de l'argent
        - Limite : Dépend du taux d'actualisation
        """)
        
        st.markdown("""
        <div class="formula-box">
        <strong>Formule VAN :</strong><br>
        VAN = -I + Σ [Ft / (1 + i)ᵗ]<br>
        Où :<br>
        I = Investissement initial<br>
        Ft = Flux de trésorerie année t<br>
        i = Taux d'actualisation<br>
        t = Période (année)
        </div>
        """, unsafe_allow_html=True)
    
    with tab2:
        st.write("### Calculateur de VAN")
        
        # Saisie des paramètres
        col_van1, col_van2 = st.columns(2)
        
        with col_van1:
            investissement = st.number_input("Investissement initial (€)", min_value=0, value=100000)
            taux_actualisation = st.slider("Taux d'actualisation (%)", 1.0, 20.0, 10.0) / 100
        
        with col_van2:
            duree = st.slider("Durée du projet (années)", 1, 10, 5)
            flux_constants = st.checkbox("Flux constants", value=True)
        
        # Saisie des flux
        st.write("**Flux de trésorerie annuels :**")
        
        flux = []
        if flux_constants:
            flux_constant = st.number_input("Flux annuel constant (€)", value=30000)
            flux = [flux_constant] * duree
        else:
            cols = st.columns(min(duree, 5))
            for i in range(duree):
                with cols[i % 5]:
                    flux_annee = st.number_input(f"Année {i+1} (€)", value=30000, key=f"flux_{i}")
                    flux.append(flux_annee)
        
        if st.button("📊 Calculer la VAN"):
            # Calcul de la VAN
            van = -investissement
            for annee in range(duree):
                flux_actualise = flux[annee] / ((1 + taux_actualisation) ** (annee + 1))
                van += flux_actualise
            
            # Affichage des résultats
            st.success(f"**VAN = {van:,.0f} €**")
            
            if van > 0:
                st.success("✅ Le projet est rentable (VAN > 0)")
            else:
                st.error("❌ Le projet n'est pas rentable (VAN ≤ 0)")
            
            # Graphique des flux actualisés
            annees = list(range(1, duree + 1))
            flux_actualises = [flux[i] / ((1 + taux_actualisation) ** (i + 1)) for i in range(duree)]
            
            fig, ax = plt.subplots(figsize=(10, 6))
            bars = ax.bar(annees, flux_actualises, color='skyblue', alpha=0.7, label='Flux actualisés')
            ax.axhline(y=0, color='black', linewidth=0.5)
            ax.set_xlabel('Année')
            ax.set_ylabel('Flux actualisés (€)')
            ax.set_title('Flux de Trésorerie Actualisés')
            ax.legend()
            
            # Ajouter les valeurs sur les barres
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                        f'{height:,.0f}€',
                        ha='center', va='bottom' if height >= 0 else 'top')
            
            st.pyplot(fig)
    
    with tab3:
        st.markdown('<div class="exercise-box">🔍 Exercice : Comparaison de Projets</div>', unsafe_allow_html=True)
        
        st.write("""
        **Énoncé :**
        Une entreprise hésite entre deux projets d'investissement. 
        Calculez la VAN de chaque projet et recommandez le meilleur.
        """)
        
        col_proj1, col_proj2 = st.columns(2)
        
        with col_proj1:
            st.write("**Projet A :**")
            invest_a = st.number_input("Investissement A (€)", value=150000, key="inv_a")
            flux_a = []
            for i in range(5):
                flux = st.number_input(f"Flux A année {i+1} (€)", value=50000, key=f"flux_a_{i}")
                flux_a.append(flux)
        
        with col_proj2:
            st.write("**Projet B :**")
            invest_b = st.number_input("Investissement B (€)", value=120000, key="inv_b")
            flux_b = []
            for i in range(5):
                flux = st.number_input(f"Flux B année {i+1} (€)", value=40000, key=f"flux_b_{i}")
                flux_b.append(flux)
        
        taux_comparaison = st.slider("Taux d'actualisation comparaison (%)", 1.0, 20.0, 10.0) / 100
        
        if st.button("📈 Comparer les projets"):
            # Calcul VAN A
            van_a = -invest_a
            for i, flux in enumerate(flux_a):
                van_a += flux / ((1 + taux_comparaison) ** (i + 1))
            
            # Calcul VAN B
            van_b = -invest_b
            for i, flux in enumerate(flux_b):
                van_b += flux / ((1 + taux_comparaison) ** (i + 1))
            
            # Affichage résultats
            col_res1, col_res2 = st.columns(2)
            with col_res1:
                st.metric("VAN Projet A", f"{van_a:,.0f} €")
            with col_res2:
                st.metric("VAN Projet B", f"{van_b:,.0f} €")
            
            if van_a > van_b and van_a > 0:
                st.success("✅ **Recommandation : Choisir le Projet A**")
            elif van_b > van_a and van_b > 0:
                st.success("✅ **Recommandation : Choisir le Projet B**")
            elif van_a <= 0 and van_b <= 0:
                st.error("❌ **Aucun projet n'est rentable**")
            else:
                st.warning("⚠️ **Les deux projets sont proches, analyse complémentaire nécessaire**")

# Section Budget de Trésorerie
def budget_tresorerie():
    st.markdown('<div class="section-header">💳 Budget de Trésorerie</div>', unsafe_allow_html=True)
    
    st.write("""
    ### Objectifs du Budget de Trésorerie
    
    - 💰 **Vérifier l'équilibre financier** mensuel
    - 📊 **Anticiper les besoins** de financement
    - 🔍 **S'assurer de la cohérence** entre tous les budgets
    - 🎯 **Prévoir la trésorerie** et éviter les découverts
    """)
    
    # Exemple interactif de budget de trésorerie
    st.subheader("📋 Modèle de Budget de Trésorerie")
    
    # Paramètres initiaux
    col_tr1, col_tr2, col_tr3 = st.columns(3)
    
    with col_tr1:
        treso_initiale = st.number_input("Trésorerie initiale (€)", value=50000)
        ca_janvier = st.number_input("CA Janvier (€)", value=80000)
        ca_fevrier = st.number_input("CA Février (€)", value=90000)
    
    with col_tr2:
        ca_mars = st.number_input("CA Mars (€)", value=100000)
        delai_clients = st.slider("Délai paiement clients (mois)", 1, 3, 2)
        taux_tva = st.slider("Taux TVA (%)", 5.5, 20.0, 20.0) / 100
    
    with col_tr3:
        charges_fixes = st.number_input("Charges fixes mensuelles (€)", value=40000)
        investissement = st.number_input("Investissement prévu (€)", value=0)
        mois_invest = st.selectbox("Mois investissement", ["Janvier", "Février", "Mars"])
    
    if st.button("📊 Générer le budget de trésorerie"):
        # Calculs simplifiés
        mois = ['Janvier', 'Février', 'Mars']
        
        # Encaissements (simplifié : 50% au comptant, 50% à 1 mois)
        encaissements = []
        for i, mois_courant in enumerate(mois):
            encaissement = 0
            if i == 0:  # Janvier
                encaissement = ca_janvier * 0.5
            elif i == 1:  # Février
                encaissement = ca_janvier * 0.5 + ca_fevrier * 0.5
            else:  # Mars
                encaissement = ca_fevrier * 0.5 + ca_mars * 0.5
            encaissements.append(encaissement)
        
        # Décaissements
        decaissements = []
        for i, mois_courant in enumerate(mois):
            decaissement = charges_fixes
            if mois_courant == mois_invest:
                decaissement += investissement
            decaissements.append(decaissement)
        
        # Calcul de la trésorerie
        tresorerie = [treso_initiale]
        for i in range(3):
            nouveau_solde = tresorerie[i] + encaissements[i] - decaissements[i]
            tresorerie.append(nouveau_solde)
        
        # Création du tableau
        data_budget = {
            'Mois': ['Janvier', 'Février', 'Mars'],
            'Encaissements': encaissements,
            'Décaissements': decaissements,
            'Trésorerie fin de mois': tresorerie[1:]
        }
        
        df_budget = pd.DataFrame(data_budget)
        
        # Affichage du tableau
        st.dataframe(df_budget.style.format({
            'Encaissements': '{:,.0f} €',
            'Décaissements': '{:,.0f} €', 
            'Trésorerie fin de mois': '{:,.0f} €'
        }), use_container_width=True)
        
        # Graphique
        fig, ax = plt.subplots(figsize=(12, 6))
        
        x = range(len(mois))
        width = 0.35
        
        ax.bar([i - width/2 for i in x], encaissements, width, label='Encaissements', color='green', alpha=0.7)
        ax.bar([i + width/2 for i in x], decaissements, width, label='Décaissements', color='red', alpha=0.7)
        ax.plot(x, tresorerie[1:], 'o-', color='blue', linewidth=3, markersize=8, label='Trésorerie')
        
        ax.set_xlabel('Mois')
        ax.set_ylabel('Montant (€)')
        ax.set_title('Budget de Trésorerie')
        ax.set_xticks(x)
        ax.set_xticklabels(mois)
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        # Ajouter les valeurs sur les barres
        for i, v in enumerate(encaissements):
            ax.text(i - width/2, v + max(encaissements)*0.01, f'{v:,.0f}€', ha='center', va='bottom')
        for i, v in enumerate(decaissements):
            ax.text(i + width/2, v + max(decaissements)*0.01, f'{v:,.0f}€', ha='center', va='bottom')
        for i, v in enumerate(tresorerie[1:]):
            ax.text(i, v + max(tresorerie[1:])*0.01, f'{v:,.0f}€', ha='center', va='bottom')
        
        st.pyplot(fig)
        
        # Analyse
        solde_min = min(tresorerie[1:])
        if solde_min < 0:
            st.error(f"⚠️ **Attention : Trésorerie négative prévue (minimum : {solde_min:,.0f} €)**")
            st.info("**Actions recommandées :** Report d'investissement, négociation de délais fournisseurs, recherche de financement")
        else:
            st.success(f"✅ **Trésorerie positive sur la période (minimum : {solde_min:,.0f} €)**")

# Section Contrôle Budgétaire
def controle_budgetaire():
    st.markdown('<div class="section-header">📊 Contrôle Budgétaire et Analyse des Écarts</div>', unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["📖 Théorie", "🧮 Calcul Écarts", "🎯 Étude de Cas"])
    
    with tab1:
        st.write("""
        ### Principe du Contrôle Budgétaire
        
        Le contrôle budgétaire consiste à comparer périodiquement les **réalisations** avec les **prévisions** 
        (budget) pour :
        
        - 📈 **Mesurer la performance**
        - 🔍 **Identifier les écarts**
        - 🎯 **Comprendre les causes**
        - 🛠️ **Proposer des actions correctives**
        
        ### Types d'Écarts
        
        **Écart = Réel - Budget**
        
        - 📊 **Écart favorable** : Réel > Budget (pour les produits) ou Réel < Budget (pour les charges)
        - 📉 **Écart défavorable** : Réel < Budget (pour les produits) ou Réel > Budget (pour les charges)
        """)
        
        st.markdown("""
        <div class="formula-box">
        <strong>Formules d'analyse des écarts :</strong><br>
        <strong>Écart sur Chiffre d'Affaires :</strong><br>
        Écart Total = CA Réel - CA Budgeté<br>
        Écart sur Prix = (Prix Réel - Prix Budgeté) × Quantité Réelle<br>
        Écart sur Quantité = (Quantité Réelle - Quantité Budgetée) × Prix Budgeté<br><br>
        
        <strong>Écart sur Coût de Production :</strong><br>
        Écart Total = Coût Réel - Coût Budgeté<br>
        Écart sur Quantité = (Quantité Réelle - Quantité Budgetée) × Prix Budgeté<br>
        Écart sur Prix = (Prix Réel - Prix Budgeté) × Quantité Réelle
        </div>
        """, unsafe_allow_html=True)
    
    with tab2:
        st.write("### Calculateur d'Écarts")
        
        col_ec1, col_ec2 = st.columns(2)
        
        with col_ec1:
            st.write("**Données budgétées :**")
            prix_budget = st.number_input("Prix budgété (€)", value=100.0)
            quantite_budget = st.number_input("Quantité budgétée", value=1000)
            cout_budget = st.number_input("Coût unitaire budgété (€)", value=60.0)
        
        with col_ec2:
            st.write("**Données réelles :**")
            prix_reel = st.number_input("Prix réel (€)", value=95.0)
            quantite_reel = st.number_input("Quantité réelle", value=900)
            cout_reel = st.number_input("Coût unitaire réel (€)", value=65.0)
        
        if st.button("📈 Analyser les écarts"):
            # Calculs
            ca_budget = prix_budget * quantite_budget
            ca_reel = prix_reel * quantite_reel
            ecart_ca_total = ca_reel - ca_budget
            
            ecart_prix = (prix_reel - prix_budget) * quantite_reel
            ecart_quantite = (quantite_reel - quantite_budget) * prix_budget
            
            cout_total_budget = cout_budget * quantite_budget
            cout_total_reel = cout_reel * quantite_reel
            ecart_cout_total = cout_total_reel - cout_total_budget
            
            ecart_cout_quantite = (quantite_reel - quantite_budget) * cout_budget
            ecart_cout_prix = (cout_reel - cout_budget) * quantite_reel
            
            # Affichage résultats
            st.subheader("📊 Résultats de l'Analyse")
            
            col_res1, col_res2 = st.columns(2)
            
            with col_res1:
                st.write("**Chiffre d'Affaires :**")
                st.metric("CA Budgeté", f"{ca_budget:,.0f} €")
                st.metric("CA Réel", f"{ca_reel:,.0f} €")
                st.metric("Écart Total", f"{ecart_ca_total:,.0f} €", 
                         delta=f"{ecart_ca_total:,.0f} €")
                
                st.write("**Décomposition :**")
                st.write(f"- Écart sur Prix : {ecart_prix:,.0f} €")
                st.write(f"- Écart sur Quantité : {ecart_quantite:,.0f} €")
            
            with col_res2:
                st.write("**Coûts de Production :**")
                st.metric("Coût Budgeté", f"{cout_total_budget:,.0f} €")
                st.metric("Coût Réel", f"{cout_total_reel:,.0f} €")
                st.metric("Écart Total", f"{ecart_cout_total:,.0f} €",
                         delta=f"{ecart_cout_total:,.0f} €")
                
                st.write("**Décomposition :**")
                st.write(f"- Écart sur Quantité : {ecart_cout_quantite:,.0f} €")
                st.write(f"- Écart sur Prix : {ecart_cout_prix:,.0f} €")
            
            # Interprétation
            st.subheader("🎯 Interprétation et Actions")
            
            if ecart_ca_total < 0:
                st.warning("**Écart défavorable sur CA**")
                if ecart_prix < 0:
                    st.write("🔍 **Cause identifiée :** Baisse des prix de vente")
                    st.write("🛠️ **Action :** Revoir la politique tarifaire")
                if ecart_quantite < 0:
                    st.write("🔍 **Cause identifiée :** Baisse des volumes vendus")
                    st.write("🛠️ **Action :** Renforcer l'effort commercial")
            
            if ecart_cout_total > 0:
                st.warning("**Écart défavorable sur coûts**")
                if ecart_cout_prix > 0:
                    st.write("🔍 **Cause identifiée :** Augmentation du coût des matières")
                    st.write("🛠️ **Action :** Renégocier avec les fournisseurs")
                if ecart_cout_quantite > 0:
                    st.write("🔍 **Cause identifiée :** Surconsommation de matières")
                    st.write("🛠️ **Action :** Améliorer les processus de production")
    
    with tab3:
        st.markdown('<div class="exercise-box">🎯 Étude de Cas : Analyse des Écarts d\'un Produit</div>', unsafe_allow_html=True)
        
        st.write("""
        **Contexte :**
        Vous êtes contrôleur de gestion dans une entreprise industrielle. 
        Analysez les écarts du produit "Gamma" pour le dernier trimestre.
        """)
        
        # Données du cas
        st.write("**Données du produit Gamma :**")
        
        col_cas1, col_cas2 = st.columns(2)
        
        with col_cas1:
            st.write("**Budget :**")
            st.write("- Prix de vente : 150 €")
            st.write("- Quantité : 2 000 unités")
            st.write("- Coût matière : 40 €/unité")
            st.write("- Quantité matière : 1.5 kg/unité")
        
        with col_cas2:
            st.write("**Réel :**")
            st.write("- Prix de vente : 145 €")
            st.write("- Quantité : 1 800 unités")
            st.write("- Coût matière : 42 €/unité")
            st.write("- Quantité matière : 1.6 kg/unité")
        
        st.write("**Questions :**")
        questions = [
            "1. Calculez l'écart total sur chiffre d'affaires",
            "2. Décomposez l'écart sur CA en écart prix et écart quantité",
            "3. Calculez l'écart total sur coût matière",
            "4. Décomposez l'écart sur coût en écart prix et écart quantité",
            "5. Proposez des actions correctives"
        ]
        
        for question in questions:
            st.write(question)
        
        with st.expander("💡 Voir la correction"):
            st.write("""
            **Correction :**
            
            1. **Écart total CA** = (145 × 1800) - (150 × 2000) = 261 000 - 300 000 = **-39 000 €** (Défavorable)
            
            2. **Décomposition CA :**
               - Écart prix = (145 - 150) × 1800 = **-9 000 €**
               - Écart quantité = (1800 - 2000) × 150 = **-30 000 €**
            
            3. **Écart total coût matière** = (42 × 1.6 × 1800) - (40 × 1.5 × 2000) = 120 960 - 120 000 = **+960 €** (Défavorable)
            
            4. **Décomposition coût :**
               - Écart prix matière = (42 - 40) × (1.6 × 1800) = **+5 760 €**
               - Écart quantité matière = (1.6 - 1.5) × 40 × 1800 = **+7 200 €**
               - Écart rendement = (1800 - 2000) × (1.5 × 40) = **-12 000 €**
            
            5. **Actions correctives :**
               - Commercial : Relancer la demande (promotions, marketing)
               - Production : Améliorer le rendement matière
               - Achats : Renégocier le prix des matières
            """)

# Section Études de Cas
def etudes_cas():
    st.markdown('<div class="section-header">🎯 Études de Cas et Exercices Globaux</div>', unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["🏭 Cas Industriel", "💼 Cas Services", "🧩 Exercices Divers"])
    
    with tab1:
        st.markdown('<div class="exercise-box">🏭 Étude de Cas : Entreprise Industrielle "MetalPro"</div>', unsafe_allow_html=True)
        
        st.write("""
        **Contexte :**
        MetalPro est une PME industrielle spécialisée dans la fabrication de pièces métalliques.
        L'entreprise rencontre des difficultés de rentabilité et souhaite revoir son pilotage.
        
        **Données :**
        - Chiffre d'affaires 2023 : 1 200 000 €
        - Résultat net 2023 : 45 000 € (3.75% du CA)
        - 2 produits principaux : P1 (60% du CA) et P2 (40% du CA)
        - Capacité de production : 10 000 heures/mois
        """)
        
        with st.expander("📋 Questions détaillées"):
            st.write("""
            **1. Analyse de la rentabilité :**
            - Calculez la marge sur coût variable par produit
            - Identifiez les produits les plus rentables
            
            **2. Optimisation de la production :**
            - P1 : Marge = 80€/unité, Temps production = 2h
            - P2 : Marge = 120€/unité, Temps production = 3h
            - Déterminez la combinaison optimale de production
            
            **3. Budget des investissements :**
            - Un nouvel équipement (200 000 €) permettrait de réduire de 20% le temps de production
            - Calculez la VAN sur 5 ans (taux 10%)
            
            **4. Tableau de bord :**
            - Proposez 5 indicateurs clés pour piloter l'entreprise
            """)
        
        with st.expander("💡 Proposition de solution"):
            st.write("""
            **Solution :**
            
            1. **Analyse de rentabilité :**
               - P1 : Marge/unité = 80€, Marge/heure = 40€
               - P2 : Marge/unité = 120€, Marge/heure = 40€
               - Même rentabilité horaire
            
            2. **Optimisation production :**
               - Fonction objectif : Max Z = 80P1 + 120P2
               - Contrainte : 2P1 + 3P2 ≤ 10 000
               - Solution : Produire selon la demande
            
            3. **Investissement :**
               - Gain temps : 20% × 10 000h = 2 000h/mois
               - Gain marge : 2 000h × 40€/h = 80 000€/mois
               - VAN positive → Investissement rentable
            
            4. **Tableau de bord :**
               - Taux de charge des machines
               - Marge sur coût variable
               - Trésorerie
               - Délai de livraison
               - Taux de rebut
            """)
    
    with tab2:
        st.markdown('<div class="exercise-box">💼 Étude de Cas : Cabinet de Conseil "StratPlus"</div>', unsafe_allow_html=True)
        
        st.write("""
        **Contexte :**
        StratPlus est un cabinet de conseil en stratégie de 20 consultants.
        L'entreprise souhaite améliorer sa profitabilité et son pilotage.
        
        **Données :**
        - CA 2023 : 2 500 000 €
        - Taux horaire moyen : 150 €
        - Taux de réalisation : 70% (temps facturable/temps travaillé)
        - Charges fixes : 800 000 €/an
        """)
        
        with st.expander("📋 Questions détaillées"):
            st.write("""
            **1. Analyse de la productivité :**
            - Calculez le nombre d'heures facturables nécessaires
            - Évaluez la marge sur coût variable
            
            **2. Budget des prestations :**
            - Établissez un budget pour 3 types de missions
            - Calculez le point mort
            
            **3. Contrôle de gestion :**
            - Proposez un système de reporting pour les consultants
            - Indicateurs de performance clés
            
            **4. Stratégie tarifaire :**
            - Analysez l'impact d'une augmentation de 10% des tarifs
            """)
    
    with tab3:
        st.write("### 🧩 Exercices Divers et QCM")
        
        # QCM
        st.subheader("📝 QCM de Validation")
        
        qcm_questions = {
            "Quelle est la formule du lot économique de Wilson ?": {
                "options": [
                    "Q* = √(2D/CL)",
                    "Q* = √(2D×CL/p)",
                    "Q* = D/CL", 
                    "Q* = CL/D"
                ],
                "correct": 1,
                "explication": "La formule correcte est Q* = √(2 × D × CL / p) où D=consommation, CL=coût lancement, p=coût possession"
            },
            "Un écart défavorable sur chiffre d'affaires peut être dû à :": {
                "options": [
                    "Une augmentation des prix",
                    "Une baisse des quantités vendues",
                    "Une diminution des coûts",
                    "Une amélioration de la productivité"
                ],
                "correct": 1,
                "explication": "Une baisse des quantités vendues réduit le chiffre d'affaires, créant un écart défavorable"
            },
            "La VAN d'un projet est positive si :": {
                "options": [
                    "Le délai de récupération est court",
                    "Les flux actualisés sont supérieurs à l'investissement",
                    "Le taux de rentabilité est élevé",
                    "Le projet dure longtemps"
                ],
                "correct": 1,
                "explication": "La VAN = Σ(flux actualisés) - investissement. Si positive, le projet crée de la valeur"
            }
        }
        
        score_qcm = 0
        user_answers_qcm = {}
        
        for i, (question, data) in enumerate(qcm_questions.items()):
            st.write(f"**{i+1}. {question}**")
            user_answers_qcm[i] = st.radio(f"Q{i}", data["options"], key=f"qcm_{i}", label_visibility="collapsed")
            
            if st.button(f"Vérifier Q{i+1}", key=f"btn_{i}"):
                if user_answers_qcm[i] == data["options"][data["correct"]]:
                    st.success("✅ Correct ! " + data["explication"])
                    score_qcm += 1
                else:
                    st.error("❌ Incorrect. " + data["explication"])
        
        if st.button("📊 Score final QCM"):
            st.info(f"Score : {score_qcm}/{len(qcm_questions)}")
            if score_qcm == len(qcm_questions):
                st.balloons()
                st.success("🎉 Excellent ! Vous maîtrisez les concepts clés du contrôle de gestion.")

# Section Applications Pratiques
def applications_pratiques():
    st.markdown('<div class="section-header">🤖 Applications Pratiques Intégrées</div>', unsafe_allow_html=True)
    
    # Application 1: Prédiction des ventes avec analyse budgétaire
    st.subheader("1. Prédiction des Ventes & Planification Budgétaire")
    
    # Génération de données historiques
    dates = pd.date_range('2023-01-01', periods=12, freq='M')
    ventes_historiques = np.random.normal(50000, 10000, 12).cumsum() + 100000
    couts_historiques = ventes_historiques * 0.6 + np.random.normal(0, 5000, 12)
    
    df_historique = pd.DataFrame({
        'Mois': dates,
        'Ventes': ventes_historiques,
        'Coûts': couts_historiques,
        'Marge': ventes_historiques - couts_historiques
    })
    
    # Modèle de prédiction
    # Préparation des données pour la prédiction
    X = np.array(range(len(df_historique))).reshape(-1, 1)
    y_ventes = df_historique['Ventes'].values
    
    model_ventes = LinearRegression()
    model_ventes.fit(X, y_ventes)
    
    # Prédiction pour les 6 prochains mois
    future_months = np.array(range(len(df_historique), len(df_historique) + 6)).reshape(-1, 1)
    ventes_predites = model_ventes.predict(future_months)
    couts_predits = ventes_predites * 0.65  # Hypothèse: coûts à 65% des ventes
    
    st.write("**Prévisions des 6 prochains mois:**")
    df_previsions = pd.DataFrame({
        'Période': ['M+1', 'M+2', 'M+3', 'M+4', 'M+5', 'M+6'],
        'Ventes Prévues': ventes_predites,
        'Coûts Prévus': couts_predits,
        'Marge Prévue': ventes_predites - couts_predits
    })
    
    # CORRECTION : Formatage spécifique par colonne
    st.dataframe(df_previsions.style.format({
        'Ventes Prévues': '{:,.2f}',
        'Coûts Prévus': '{:,.2f}', 
        'Marge Prévue': '{:,.2f}'
    }))
    
    # Visualisation des prévisions
    st.subheader("📈 Visualisation des Prévisions")
    
    # Préparation des données pour le graphique
    dates_futures = pd.date_range(df_historique['Mois'].iloc[-1] + pd.Timedelta(days=31), periods=6, freq='M')
    
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # Données historiques
    ax.plot(df_historique['Mois'], df_historique['Ventes'], 'b-', label='Ventes Historiques', marker='o')
    ax.plot(df_historique['Mois'], df_historique['Coûts'], 'r-', label='Coûts Historiques', marker='s')
    
    # Prévisions
    ax.plot(dates_futures, ventes_predites, 'b--', label='Ventes Prévues', marker='o')
    ax.plot(dates_futures, couts_predits, 'r--', label='Coûts Prévus', marker='s')
    
    ax.set_title('Prévisions des Ventes et Coûts - Modèle de Régression Linéaire')
    ax.set_xlabel('Mois')
    ax.set_ylabel('Montant (€)')
    ax.legend()
    ax.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    st.pyplot(fig)
    
    # Métriques de performance du modèle
    st.subheader("📊 Performance du Modèle")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Calcul R²
        predictions_historiques = model_ventes.predict(X)
        r2 = r2_score(y_ventes, predictions_historiques)
        st.metric("Score R² du Modèle", f"{r2:.3f}")
    
    with col2:
        # Erreur moyenne
        mae = mean_absolute_error(y_ventes, predictions_historiques)
        st.metric("Erreur Moyenne Absolue", f"{mae:.2f} €")
    
    with col3:
        # Croissance moyenne prévue
        croissance_moyenne = ((ventes_predites[-1] - ventes_predites[0]) / ventes_predites[0]) * 100
        st.metric("Croissance Moyenne Prévue", f"{croissance_moyenne:.1f}%")
    
    # Application 2: Analyse de sensibilité
    st.subheader("2. Analyse de Sensibilité Budgétaire")
    
    st.write("**Simulation de différents scénarios:**")
    
    col1, col2 = st.columns(2)
    
    with col1:
        taux_croissance = st.slider("Taux de croissance (%)", -10.0, 20.0, 5.0, 0.5)
        ratio_couts = st.slider("Ratio coûts/ventes (%)", 50.0, 80.0, 65.0, 1.0)
    
    with col2:
        inflation = st.slider("Taux d'inflation (%)", 0.0, 10.0, 2.0, 0.1)
        investissement_marketing = st.number_input("Investissement marketing (k€)", 0, 100, 10)
    
    # Simulation avec les nouveaux paramètres
    ventes_simulees = ventes_predites * (1 + taux_croissance/100) + investissement_marketing * 1000
    couts_simules = ventes_simulees * (ratio_couts/100) * (1 + inflation/100)
    marges_simulees = ventes_simulees - couts_simules
    
    df_simulation = pd.DataFrame({
        'Période': ['M+1', 'M+2', 'M+3', 'M+4', 'M+5', 'M+6'],
        'Ventes Simulées': ventes_simulees,
        'Coûts Simulés': couts_simules,
        'Marge Simulée': marges_simulees
    })
    
    # CORRECTION : Formatage spécifique par colonne
    st.dataframe(df_simulation.style.format({
        'Ventes Simulées': '{:,.2f}',
        'Coûts Simulés': '{:,.2f}',
        'Marge Simulée': '{:,.2f}'
    }))
    
    # Application 3: Recommandations stratégiques
    st.subheader("3. Recommandations Stratégiques Automatisées")
    
    recommendations = []
    
    # Analyse de la rentabilité
    marge_moyenne = np.mean(marges_simulees)
    if marge_moyenne < 10000:
        recommendations.append("🔴 **Optimisation des coûts nécessaire** - La marge moyenne est inférieure à 10k€")
    elif marge_moyenne > 30000:
        recommendations.append("🟢 **Excellente rentabilité** - Poursuivre la stratégie actuelle")
    else:
        recommendations.append("🟡 **Rentabilité correcte** - Opportunités d'amélioration identifiées")
    
    # Analyse de la croissance
    croissance_simulee = ((ventes_simulees[-1] - ventes_simulees[0]) / ventes_simulees[0]) * 100
    if croissance_simulee > 15:
        recommendations.append("🚀 **Forte croissance** - Anticiper les besoins en capacité")
    elif croissance_simulee < 0:
        recommendations.append("📉 **Déclin des ventes** - Revoyer la stratégie commerciale")
    else:
        recommendations.append("📈 **Croissance modérée** - Maintenir les efforts actuels")
    
    # Analyse de la stabilité
    volatilite_ventes = np.std(ventes_simulees) / np.mean(ventes_simulees) * 100
    if volatilite_ventes > 20:
        recommendations.append("⚡ **Forte volatilité** - Renforcer la prévision des ventes")
    
    # Affichage des recommandations
    st.write("**Recommandations basées sur l'analyse des données:**")
    for i, recommendation in enumerate(recommendations, 1):
        st.write(f"{i}. {recommendation}")
    
    # Application 4: Dashboard synthétique
    st.subheader("4. Tableau de Bord Synthétique")
    
    # Création d'un dashboard avec métriques clés
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "CA Annuel Prévisionnel", 
            f"{(ventes_simulees.sum() / 1000):.0f}K €",
            delta=f"{(ventes_simulees.sum() - ventes_predites.sum()) / 1000:.0f}K €"
        )
    
    with col2:
        st.metric(
            "Marge Moyenne", 
            f"{(marges_simulees.mean() / 1000):.1f}K €",
            delta=f"{(marges_simulees.mean() - (ventes_predites - couts_predits).mean()) / 1000:.1f}K €"
        )
    
    with col3:
        st.metric(
            "Taux de Marge", 
            f"{(marges_simulees.mean() / ventes_simulees.mean() * 100):.1f}%"
        )
    
    with col4:
        st.metric(
            "ROI Marketing", 
            f"{((ventes_simulees.sum() - ventes_predites.sum()) / (investissement_marketing * 1000)):.1f}x"
        )
    
    # Application 5: Analyse de clustering des clients
    st.subheader("5. Segmentation Clients par Intelligence Artificielle")
    
    # Génération de données clients simulées
    np.random.seed(42)
    n_clients = 200
    
    data_clients = {
        'CA_Annuel': np.random.gamma(2, 25000, n_clients),
        'Frequence_Achat': np.random.poisson(12, n_clients),
        'Marge_Client': np.random.normal(15000, 5000, n_clients),
        'Anciennete': np.random.exponential(3, n_clients) + 1
    }
    
    df_clients = pd.DataFrame(data_clients)
    
    # Clustering avec K-Means
    scaler = StandardScaler()
    features = ['CA_Annuel', 'Frequence_Achat', 'Marge_Client', 'Anciennete']
    X_scaled = scaler.fit_transform(df_clients[features])
    
    kmeans = KMeans(n_clusters=4, random_state=42)
    df_clients['Segment_IA'] = kmeans.fit_predict(X_scaled)
    
    # Réduction de dimension pour visualisation
    pca = PCA(n_components=2)
    X_pca = pca.fit_transform(X_scaled)
    df_clients['PCA1'] = X_pca[:, 0]
    df_clients['PCA2'] = X_pca[:, 1]
    
    col_viz1, col_viz2 = st.columns(2)
    
    with col_viz1:
        # Visualisation des segments
        fig = px.scatter(
            df_clients, 
            x='PCA1', 
            y='PCA2',
            color='Segment_IA',
            title='👥 Segmentation Clients - Analyse PCA',
            hover_data=['CA_Annuel', 'Marge_Client'],
            color_continuous_scale='viridis'
        )
        
        # CORRECTION : Ajout d'une clé unique
        st.plotly_chart(fig, use_container_width=True, key="segmentation_clients")
    
    with col_viz2:
        # Profils des segments
        analyse_segments = df_clients.groupby('Segment_IA').agg({
            'CA_Annuel': 'mean',
            'Marge_Client': 'mean',
            'Frequence_Achat': 'mean',
            'Anciennete': 'mean'
        }).round(2)
        
        st.write("**📊 Profils des Segments Identifiés**")
        st.dataframe(analyse_segments.style.format("{:.0f}"))
        
        # Recommandations par segment
        st.write("**🎯 Recommandations Stratégiques**")
        
        recommandations_segments = {
            0: "💎 **Clients Premium** - Développer services exclusifs",
            1: "📈 **Clients Croissance** - Programmes de fidélisation", 
            2: "🔄 **Clients Réguliers** - Cross-selling stratégique",
            3: "⚠️ **Clients à Risque** - Actions de rétention"
        }
        
        for segment, recommandation in recommandations_segments.items():
            st.info(recommandation)

# Section Ressources
def ressources():
    st.markdown('<div class="section-header">📖 Ressources et Approfondissements</div>', unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["📚 Bibliographie", "📋 Glossaire", "🔧 Templates"])
    
    with tab1:
        st.write("### 📚 Bibliographie Recommandée")
        
        ressources = {
            "Ouvrages Fondamentaux": [
                "**Contrôle de gestion et gestion budgétaire** - Alain Mikol",
                "**Le Contrôle de gestion** - François Dupuy (Que Sais-Je ?)",
                "**Pratique du contrôle de gestion** - Pierre-Laurent Bescos"
            ],
            "Ouvrages Avancés": [
                "**Tableaux de bord et pilotage de la performance** - Alain Fernandez",
                "**Contrôle de gestion dans les services** - Jean-Guy Degos",
                "**La méthode ABC/ABM** - Robert Kaplan"
            ],
            "Ressources en Ligne": [
                "[AFIGE - Association Francophone de Comptabilité et Gestion](https://www.afige.com)",
                "[Dauphine - Cours en ligne contrôle de gestion](https://dauphine.psl.eu)",
                "[Village de la Gestion - Ressources gratuites](https://www.village-gestion.com)"
            ]
        }
        
        for categorie, items in ressources.items():
            st.write(f"#### {categorie}")
            for item in items:
                st.write(f"- {item}")
    
    with tab2:
        st.write("### 📋 Glossaire des Termes Techniques")
        
        glossaire = {
            "📊 Budget": "Plan chiffré des recettes et dépenses prévues pour une période donnée",
            "🎯 Écart": "Différence entre une réalisation et une prévision budgétaire",
            "💰 VAN (Valeur Actuelle Nette)": "Différence entre la valeur actuelle des flux futurs et l'investissement initial",
            "📦 Lot Économique": "Quantité optimale à commander pour minimiser les coûts de stock",
            "🏭 Marge sur Coût Variable": "Différence entre le prix de vente et les coûts variables",
            "📈 Point Mort": "Niveau d'activité pour lequel l'entreprise couvre tous ses coûts",
            "🔍 Tableau de Bord": "Outil de pilotage présentant les indicateurs clés de performance",
            "💳 Trésorerie": "Ensemble des disponibilités monétaires de l'entreprise"
        }
        
        for terme, definition in glossaire.items():
            st.write(f"**{terme}** : {definition}")
    
    with tab3:
        st.write("### 🔧 Templates à Télécharger")
        
        st.write("""
        Modèles Excel prêts à l'emploi pour votre contrôle de gestion :
        """)
        
        # Template 1: Budget des Ventes
        def create_budget_ventes():
            """Crée un template Excel pour le budget des ventes"""
            output = BytesIO()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Budget des Ventes"
            
            # En-têtes
            headers = ['Mois', 'Quantité', 'Prix Unitaire', 'Chiffre d\'Affaires', 'Objectif', 'Écart']
            sheet.append(headers)
            
            # Données exemple
            mois = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 
                   'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
            
            for i, mois in enumerate(mois):
                quantite = 1000 + i * 50
                prix = 150
                ca = quantite * prix
                objectif = 160000
                ecart = ca - objectif
                
                sheet.append([mois, quantite, prix, ca, objectif, ecart])
            
            # Formules et mise en forme
            for row in range(2, 14):
                sheet[f'D{row}'] = f'=B{row}*C{row}'
                sheet[f'F{row}'] = f'=D{row}-E{row}'
            
            # Total
            sheet.append(['TOTAL', f'=SUM(B2:B13)', '', f'=SUM(D2:D13)', f'=SUM(E2:E13)', f'=SUM(F2:F13)'])
            
            workbook.save(output)
            return output.getvalue()
        
        # Template 2: Budget de Production
        def create_budget_production():
            """Crée un template Excel pour le budget de production"""
            output = BytesIO()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Budget Production"
            
            headers = ['Produit', 'Quantité à Produire', 'Temps Unitaire (h)', 'Temps Total (h)', 
                      'Coût Matière/unité', 'Coût Main d\'œuvre/h', 'Coût Total']
            sheet.append(headers)
            
            produits = ['Produit A', 'Produit B', 'Produit C']
            for produit in produits:
                qte = 500
                temps_unit = 2
                temps_total = qte * temps_unit
                cout_matiere = 25
                cout_mo = 45
                cout_total = (cout_matiere + cout_mo * temps_unit) * qte
                
                sheet.append([produit, qte, temps_unit, temps_total, cout_matiere, cout_mo, cout_total])
            
            # Total
            sheet.append(['TOTAL', f'=SUM(B2:B4)', '', f'=SUM(D2:D4)', '', '', f'=SUM(G2:G4)'])
            
            workbook.save(output)
            return output.getvalue()
        
        # Template 3: Gestion des Stocks
        def create_gestion_stocks():
            """Crée un template Excel pour la gestion des stocks"""
            output = BytesIO()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Wilson et Stocks"
            
            # Section paramètres
            sheet.append(['PARAMÈTRES'])
            sheet.append(['Consommation annuelle (D)', 10000])
            sheet.append(['Coût de lancement (CL)', 150])
            sheet.append(['Prix unitaire', 25])
            sheet.append(['Taux de possession (%)', 15])
            sheet.append(['Coût possession unitaire (p)', '=D3*D5/100'])
            
            # Section résultats
            sheet.append([])
            sheet.append(['RÉSULTATS'])
            sheet.append(['Lot économique (Q*)', '=SQRT(2*D2*D3/D6)'])
            sheet.append(['Nombre de commandes/an', '=D2/D9'])
            sheet.append(['Période économique (jours)', '=365/D10'])
            
            # Section suivi stocks
            sheet.append([])
            sheet.append(['SUIVI DES STOCKS'])
            sheet.append(['Mois', 'Stock Début', 'Entrées', 'Sorties', 'Stock Fin'])
            mois = ['Janvier', 'Février', 'Mars']
            for mois in mois:
                sheet.append([mois, 1000, 2000, 1500, '=B15+C15-D15'])
            
            workbook.save(output)
            return output.getvalue()
        
        # Template 4: Budget d'Investissement
        def create_budget_investissement():
            """Crée un template Excel pour le budget d'investissement"""
            output = BytesIO()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Analyse Investissement"
            
            # En-têtes VAN
            headers = ['Année', 'Investissement', 'Flux de Trésorerie', 'Taux Actualisation', 'Flux Actualisé']
            sheet.append(headers)
            
            # Données exemple
            investissement = -100000
            taux = 0.1
            flux = [30000, 35000, 40000, 45000, 50000]
            
            sheet.append([0, investissement, '', taux, '=C2'])
            for i, flux_annee in enumerate(flux, 1):
                flux_actualise = f'=C{i+2}/(1+D$2)^{i}'
                sheet.append([i, '', flux_annee, '', flux_actualise])
            
            # VAN
            sheet.append(['', '', '', 'VAN', f'=SUM(E2:E7)'])
            
            # Section critères
            sheet.append([])
            sheet.append(['CRITÈRES DE DÉCISION'])
            sheet.append(['VAN', '=E8'])
            sheet.append(['Délai de récupération', 'À calculer manuellement'])
            sheet.append(['Taux de rentabilité interne', 'À calculer avec TRI()'])
            
            workbook.save(output)
            return output.getvalue()
        
        # Template 5: Budget de Trésorerie
        def create_budget_tresorerie():
            """Crée un template Excel pour le budget de trésorerie"""
            output = BytesIO()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Budget Trésorerie"
            
            headers = ['Mois', 'Trésorerie Début', 'Encaissements', 'Décaissements', 'Trésorerie Fin']
            sheet.append(headers)
            
            mois = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin']
            treso_debut = 50000
            
            for i, mois in enumerate(mois):
                if i == 0:
                    treso_debut_mois = treso_debut
                else:
                    treso_debut_mois = f'=E{2+i}'
                
                encaissements = 80000
                decaissements = 75000
                treso_fin = f'=B{2+i}+C{2+i}-D{2+i}'
                
                sheet.append([mois, treso_debut_mois, encaissements, decaissements, treso_fin])
            
            workbook.save(output)
            return output.getvalue()
        
        # Template 6: Tableau de Bord
        def create_tableau_bord():
            """Crée un template Excel pour le tableau de bord"""
            output = BytesIO()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Tableau de Bord"
            
            # Indicateurs clés
            sheet.append(['INDICATEURS CLÉS DE PERFORMANCE'])
            sheet.append(['Indicateur', 'Valeur', 'Cible', 'Écart', 'Tendance'])
            
            indicateurs = [
                ['Chiffre d\'affaires', 250000, 240000, '=B3-C3', '✅'],
                ['Marge commerciale', 62500, 60000, '=B4-C4', '✅'],
                ['Taux de rentabilité', '=B4/B3', '0.25', '=B5-C5', '✅'],
                ['Point mort (CA)', 192000, 180000, '=B6-C6', '⚠️'],
                ['Ratio fonds de roulement', 1.2, 1.5, '=B7-C7', '❌'],
                ['Délai clients (jours)', 45, 30, '=B8-C8', '❌']
            ]
            
            for indicateur in indicateurs:
                sheet.append(indicateur)
            
            # Graphique données
            sheet.append([])
            sheet.append(['SUIVI MENSUEL'])
            sheet.append(['Mois', 'CA Réel', 'CA Objectif'])
            mois_ca = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin']
            for i, mois in enumerate(mois_ca):
                ca_reel = 20000 + i * 5000
                ca_objectif = 19000 + i * 5500
                sheet.append([mois, ca_reel, ca_objectif])
            
            workbook.save(output)
            return output.getvalue()
        
        # Affichage des templates avec boutons de téléchargement
        templates_data = {
            "📊 Budget des Ventes": {
                "function": create_budget_ventes,
                "description": "Template pour établir et suivre le budget des ventes avec calcul automatique des écarts"
            },
            "🏭 Budget de Production": {
                "function": create_budget_production,
                "description": "Modèle d'optimisation de la production avec calcul des coûts et temps"
            },
            "📦 Gestion des Stocks": {
                "function": create_gestion_stocks,
                "description": "Calculateur de lot économique (Wilson) et suivi des stocks avec formules intégrées"
            },
            "💰 Budget d'Investissement": {
                "function": create_budget_investissement,
                "description": "Calculateur de VAN et analyse de rentabilité des projets d'investissement"
            },
            "💳 Budget de Trésorerie": {
                "function": create_budget_tresorerie,
                "description": "Tableau de flux de trésorerie mensuel avec calcul automatique du solde"
            },
            "🎯 Tableau de Bord": {
                "function": create_tableau_bord,
                "description": "Modèle de tableau de bord avec indicateurs clés de performance et tendances"
            }
        }
        
        for template_name, template_info in templates_data.items():
            col_t1, col_t2 = st.columns([3, 1])
            with col_t1:
                st.write(f"**{template_name}**")
                st.write(template_info["description"])
            with col_t2:
                # Générer le fichier Excel
                excel_data = template_info["function"]()
                
                st.download_button(
                    label="📥 Télécharger",
                    data=excel_data,
                    file_name=f"{template_name.replace(' ', '_').replace('📊', '').replace('🏭', '').replace('📦', '').replace('💰', '').replace('💳', '').replace('🎯', '').strip()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_{template_name}"
                )
        
        st.info("""
        💡 **Conseil d'utilisation :**
        - Les templates contiennent des formules Excel pré-configurées
        - Personnalisez les données d'exemple avec vos propres chiffres
        - Les calculs se mettent à jour automatiquement
        - Idéal pour démarrer rapidement votre contrôle de gestion
        """)

# Footer de l'application
def footer():
    st.markdown("---")
    
    # Footer principal
    st.markdown("""
    <div class="footer">
        <h3>🎯 Résumé de l'Application</h3>
        <p><strong>Contrôle de Gestion - Application Interactive</strong> est une plateforme complète d'apprentissage 
        et de pratique du contrôle de gestion moderne, intégrant les méthodes traditionnelles avec les outils 
        de Data Science.</p>
        
        
    """, unsafe_allow_html=True)
    
    # Information sur l'auteur
    st.markdown("""
    <div class="author-info">
        <h6>👨‍💻 Ibou coumba Gueye Xataxeli</h6>
        <p><strong>Expert en Contrôle de Gestion & Data Science</strong></p>
        <p>Cette application a été développée pour démocratiser l'accès aux outils modernes 
        de contrôle de gestion et faciliter l'apprentissage des concepts fondamentaux.</p>
        <p>📧 <em>Application éducative - Tous droits réservés</em></p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()